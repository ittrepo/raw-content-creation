import uuid
import random
import hashlib
import base64
from datetime import datetime, timezone, timedelta
import requests
import xml.etree.ElementTree as ET
import json
import openpyxl
from dotenv import load_dotenv
import os

load_dotenv()

def xml_to_dict(element):
    result = {}
    if element.attrib:
        result.update(element.attrib)
    if element.text and element.text.strip():
        if result:
            result['text'] = element.text.strip()
        else:
            result = element.text.strip()
    for child in element:
        child_data = xml_to_dict(child)
        if child.tag in result:
            if isinstance(result[child.tag], list):
                result[child.tag].append(child_data)
            else:
                result[child.tag] = [result[child.tag], child_data]
        else:
            result[child.tag] = child_data
    return result

def build_soap_request(latitude, longitude):
    # --- All your static parameters here ---
    organization = 'NMC-SAUDI'
    user_id = os.getenv('AMADEUSE_USER_ID')
    password = 'psC6Gh=q3qPb'
    office_id = 'DMMS228XU'
    duty_code = 'SU'
    requestor_type = 'U'
    wsap = '1ASIWAAAAAK'
    url = os.getenv('AMADEUSE_LIVE_URL')
    uuid_value = str(uuid.uuid4())
    t = datetime.now(timezone.utc)
    micro = f"{t.microsecond:06d}"[:-3]
    date_str = t.strftime("%Y-%m-%dT%H:%M:%S") + micro + 'Z'
    nonce = random.randint(10000000, 99999999)
    nounce = base64.b64encode(str(nonce).encode()).decode()
    pass_sha = base64.b64encode(hashlib.sha1(str(nonce).encode() + date_str.encode() + hashlib.sha1(password.encode()).digest()).digest()).decode()
    start_date = (datetime.now(timezone.utc) + timedelta(days=1)).strftime("%Y-%m-%d")
    end_date = (datetime.now(timezone.utc) + timedelta(days=2)).strftime("%Y-%m-%d")
    soap_xml_request = f"""
                    <soap:Envelope xmlns:soap='http://schemas.xmlsoap.org/soap/envelope/'>
                    <soap:Header>
                        <add:MessageID xmlns:add='http://www.w3.org/2005/08/addressing'>{uuid_value}</add:MessageID>
                        <add:Action xmlns:add='http://www.w3.org/2005/08/addressing'>http://webservices.amadeus.com/Hotel_MultiSingleAvailability_10.0</add:Action>
                        <add:To xmlns:add='http://www.w3.org/2005/08/addressing'>{url}/{wsap}</add:To>
                        <oas:Security xmlns:oas='http://docs.oasis-open.org/wss/2004/01/oasis-200401-wss-wssecurity-secext-1.0.xsd'>
                        <oas:UsernameToken xmlns:oas1='http://docs.oasis-open.org/wss/2004/01/oasis-200401-wss-wssecurity-utility-1.0.xsd' oas1:Id='UsernameToken-1'>
                            <oas:Username>{user_id}</oas:Username>
                            <oas:Nonce EncodingType='http://docs.oasis-open.org/wss/2004/01/oasis-200401-wss-soap-message-security-1.0#Base64Binary'>{nounce}</oas:Nonce>
                            <oas:Password Type='http://docs.oasis-open.org/wss/2004/01/oasis-200401-wss-username-token-profile-1.0#PasswordDigest'>{pass_sha}</oas:Password>
                            <oas1:Created>{date_str}</oas1:Created>
                        </oas:UsernameToken>
                        </oas:Security>
                        <AMA_SecurityHostedUser xmlns='http://xml.amadeus.com/2010/06/Security_v1'>
                        <UserID POS_Type='1' PseudoCityCode='{office_id}' AgentDutyCode='{duty_code}' RequestorType='{requestor_type}'/>
                        </AMA_SecurityHostedUser>
                    </soap:Header>
                    <soap:Body>
                    <OTA_HotelAvailRQ EchoToken='MultiSingle' Version='4.000' SummaryOnly='true' AvailRatesOnly='true' OnRequestInd='true' RateRangeOnly='true' ExactMatchOnly='false' RateDetailsInd='true' SearchCacheLevel='Live' RequestedCurrency='USD'>
                    <AvailRequestSegments>
                        <AvailRequestSegment InfoSource='Distribution'>
                            <HotelSearchCriteria>
                                <Criterion ExactMatch='true'>
                                    <Position Latitude='{latitude}' Longitude='{longitude}'></Position>
                                    <Radius Distance='50' DistanceMeasure='DIS' UnitOfMeasureCode='2'></Radius>
                                    <StayDateRange Start='{start_date}' End='{end_date}' />
                                    <RoomStayCandidates>
                                        <RoomStayCandidate RoomID='1' Quantity='1'>
                                            <GuestCounts>
                                                <GuestCount AgeQualifyingCode='10' Count='1' />
                                            </GuestCounts>
                                        </RoomStayCandidate>
                                    </RoomStayCandidates>
                                </Criterion>
                            </HotelSearchCriteria>
                        </AvailRequestSegment>
                    </AvailRequestSegments>
                    </OTA_HotelAvailRQ>
                    </soap:Body>
                    </soap:Envelope>
                    """
    return soap_xml_request, url

def extract_and_save_hotel_info(response_content, output_dir):
    root = ET.fromstring(response_content)
    response_dict = xml_to_dict(root)

    try:
        body = response_dict.get('{http://schemas.xmlsoap.org/soap/envelope/}Body', {})
        avail_rs = body.get('{http://www.opentravel.org/OTA/2003/05}OTA_HotelAvailRS', {})
        hotel_stays = avail_rs.get('{http://www.opentravel.org/OTA/2003/05}HotelStays', {})
        hotel_list = hotel_stays.get('{http://www.opentravel.org/OTA/2003/05}HotelStay', [])

        if isinstance(hotel_list, dict):
            hotel_list = [hotel_list]

        for hotel in hotel_list:
            basic_info = hotel.get('{http://www.opentravel.org/OTA/2003/05}BasicPropertyInfo', {})
            hotel_code = basic_info.get('HotelCode')

            if hotel_code:
                out_path = os.path.join(output_dir, f"{hotel_code}.json")

                # 👉 Skip if file already exists
                if os.path.exists(out_path):
                    print(f"⏩⏩⏩⏩⏩ Skipped existing hotel: {hotel_code}")
                    continue

                out_data = {
                    "RoomStayRPH": hotel.get("RoomStayRPH"),
                    "{http://www.opentravel.org/OTA/2003/05}BasicPropertyInfo": basic_info
                }

                with open(out_path, "w", encoding="utf-8") as f:
                    json.dump(out_data, f, indent=2)
                print(f"Saved {out_path}")

    except Exception as e:
        print(f"Error extracting hotel info: {e}")

def main():
    xlsx_path = "hotels_cities.xlsx"
    output_dir = r"D:\content_for_hotel_json\cdn_row_collection\amadeuse\with_location"
    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Find column indices
    lat_col = lon_col = status_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "Latitude":
            lat_col = idx
        elif cell.value == "Longitude":
            lon_col = idx
        elif cell.value == "status":
            status_col = idx

    # If "status" column doesn't exist, add it
    if status_col is None:
        status_col = ws.max_column + 1
        ws.cell(row=1, column=status_col).value = "status"

    if lat_col is None or lon_col is None:
        print("Latitude or Longitude column not found.")
        return

    for row in range(2, ws.max_row + 1):
        status_value = ws.cell(row=row, column=status_col).value
        if status_value and str(status_value).strip().lower() == "done":
            continue  

        raw_lat = ws.cell(row=row, column=lat_col).value
        raw_lon = ws.cell(row=row, column=lon_col).value

        if not isinstance(raw_lat, (float, int)) or not isinstance(raw_lon, (float, int)):
            continue

        # ←←← change multiplier from 10 000 to 100 000 here
        latitude  = int(raw_lat  * 100_000)
        longitude = int(raw_lon  * 100_000)


        soap_xml_request, url = build_soap_request(latitude, longitude)
        header_data = {
            "Content-Type": "text/xml",
            "SOAPAction": "http://webservices.amadeus.com/Hotel_MultiSingleAvailability_10.0"
        }

        try:
            response = requests.post(url=url, headers=header_data, data=soap_xml_request, timeout=30)
            if response.status_code == 200:
                extract_and_save_hotel_info(response.content, output_dir)
                ws.cell(row=row, column=status_col).value = "Done"
                wb.save(xlsx_path)
                print(f"✅ Done: {latitude}, {longitude}")
            else:
                print(f"❌ Request failed: {latitude}, {longitude} - Status: {response.status_code}")
        except Exception as ex:
            print(f"⚠️ Request error: {latitude}, {longitude} - Error: {ex}")

    # Final save
    wb.save(xlsx_path)
    print("✔️ Excel file updated.")

if __name__ == "__main__":
    main()
